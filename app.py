"""
Krishna Land Developers - CRM System
Flask + SQLite Backend
"""

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
import sqlite3
import hashlib
import os
import json
import random
import string
from datetime import datetime, timedelta
from functools import wraps
import io

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'krishna_land_dev_secret_2024_secure_key'
app.permanent_session_lifetime = timedelta(hours=8)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'krishna_crm.db')

# ─────────────────────────────────────────────
# DATABASE SETUP
# ─────────────────────────────────────────────


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT UNIQUE,
            full_name TEXT NOT NULL,
            email TEXT UNIQUE,
            mobile TEXT UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'caller',
            is_active INTEGER DEFAULT 1,
            last_login TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            mobile TEXT NOT NULL,
            whatsapp TEXT,
            email TEXT,
            city TEXT,
            state TEXT,
            age_group TEXT,
            status TEXT DEFAULT 'New',
            not_connected_reason TEXT,
            caller_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(caller_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS customer_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER UNIQUE,
            occupation TEXT,
            company_name TEXT,
            monthly_income TEXT,
            investment_experience TEXT,
            budget_range TEXT,
            payment_preference TEXT,
            loan_bank TEXT,
            loan_amount TEXT,
            loan_tenure_months INTEGER,
            purpose TEXT,
            holding_period TEXT,
            dholera_aware TEXT,
            awareness_source TEXT,
            interest_drivers TEXT,
            plot_size TEXT,
            project_type TEXT,
            risk_appetite TEXT,
            investment_timeline TEXT,
            lead_source TEXT,
            assigned_executive TEXT,
            follow_up_date TEXT,
            remarks TEXT,
            lead_score INTEGER DEFAULT 0,
            lead_category TEXT DEFAULT 'Cold',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(lead_id) REFERENCES leads(id)
        );

        CREATE TABLE IF NOT EXISTS call_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER,
            caller_id INTEGER,
            notes TEXT,
            duration_minutes INTEGER DEFAULT 0,
            call_date TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(lead_id) REFERENCES leads(id),
            FOREIGN KEY(caller_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS follow_ups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lead_id INTEGER,
            caller_id INTEGER,
            follow_up_date TEXT,
            status TEXT DEFAULT 'Pending',
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(lead_id) REFERENCES leads(id),
            FOREIGN KEY(caller_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS login_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            ip_address TEXT,
            device_info TEXT,
            login_time TEXT DEFAULT CURRENT_TIMESTAMP,
            status TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
    ''')

    # Migrations: add columns if they don't exist
    for col, ctype in [
        ('loan_bank', 'TEXT'), ('loan_amount', 'TEXT'),
        ('loan_tenure_months', 'INTEGER'), ('not_connected_reason', 'TEXT')
    ]:
        try:
            if col in ('loan_bank', 'loan_amount', 'loan_tenure_months'):
                c.execute(
                    f'ALTER TABLE customer_profiles ADD COLUMN {col} {ctype}')
            else:
                c.execute(f'ALTER TABLE leads ADD COLUMN {col} {ctype}')
        except Exception:
            pass

    # Default admin
    for emp, name, email, mobile, pwd, role in [
        ('KLD001', 'Admin User', 'admin@krishnaland.com',
         '9999999999', 'Admin@123', 'admin'),
        ('KLD002', 'Rajesh Sharma', 'rajesh@krishnaland.com',
         '9888888888', 'Manager@123', 'manager'),
        ('KLD003', 'Priya Patel', 'priya@krishnaland.com',
         '9777777777', 'Caller@123', 'caller'),
    ]:
        try:
            c.execute('''INSERT OR IGNORE INTO users
                (employee_id, full_name, email, mobile, password_hash, role)
                VALUES (?,?,?,?,?,?)''',
                      (emp, name, email, mobile, hash_password(pwd), role))
        except Exception:
            pass

    conn.commit()
    conn.close()

init_db()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def generate_employee_id(conn):
    row = conn.execute(
        "SELECT employee_id FROM users ORDER BY id DESC LIMIT 1").fetchone()
    if row and row['employee_id'] and row['employee_id'].startswith('KLD'):
        try:
            num = int(row['employee_id'][3:]) + 1
            return f"KLD{num:03d}"
        except Exception:
            pass
    count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    return f"KLD{count + 1:03d}"


# ─────────────────────────────────────────────
# AUTH DECORATORS
# ─────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to continue.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if session.get('role') not in roles:
                flash('Access denied.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator


# ─────────────────────────────────────────────
# LEAD SCORING – always live
# ─────────────────────────────────────────────

def calculate_lead_score(profile):
    score = 0
    budget = (profile.get('budget_range') or '')
    if budget in ['₹25–50 Lakhs', '₹50 Lakhs – ₹1 Cr', '₹1 Cr+']:
        score += 3
    timeline = (profile.get('investment_timeline') or '')
    if timeline in ['Immediate', 'Within 1 Month']:
        score += 3
    if (profile.get('dholera_aware') or '') == 'Yes':
        score += 2
    if (profile.get('investment_experience') or '') == 'Experienced Investor':
        score += 2
    if (profile.get('payment_preference') or '') == 'Full Payment':
        score += 2
    if (profile.get('risk_appetite') or '') in ['Moderate', 'High']:
        score += 1
    category = 'Hot' if score >= 10 else ('Warm' if score >= 6 else 'Cold')
    return score, category


def recalculate_and_save_score(conn, lead_id):
    profile = conn.execute(
        'SELECT * FROM customer_profiles WHERE lead_id=?', (lead_id,)).fetchone()
    if profile:
        score, category = calculate_lead_score(dict(profile))
        conn.execute('UPDATE customer_profiles SET lead_score=?, lead_category=? WHERE lead_id=?',
                     (score, category, lead_id))


# ─────────────────────────────────────────────
# ROUTES: AUTH
# ─────────────────────────────────────────────

@app.route('/', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        identifier = request.form.get('identifier', '').strip()
        password = request.form.get('password', '').strip()
        captcha_input = request.form.get('captcha', '').strip()
        if captcha_input.upper() != session.get('captcha', '').upper():
            error = 'Invalid CAPTCHA. Please try again.'
        else:
            conn = get_db()
            user = conn.execute(
                'SELECT * FROM users WHERE (employee_id=? OR email=? OR mobile=?) AND is_active=1',
                (identifier, identifier, identifier)).fetchone()
            conn.close()
            if user and user['password_hash'] == hash_password(password):
                session.permanent = 'remember' in request.form
                session['user_id'] = user['id']
                session['username'] = user['full_name']
                session['role'] = user['role']
                session['employee_id'] = user['employee_id']
                conn = get_db()
                conn.execute('INSERT INTO login_logs (user_id, ip_address, device_info, status) VALUES (?,?,?,?)',
                             (user['id'], request.remote_addr,
                              request.headers.get('User-Agent', '')[:200], 'success'))
                conn.execute('UPDATE users SET last_login=? WHERE id=?',
                             (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
                conn.commit()
                conn.close()
                flash(f'Welcome back, {user["full_name"]}!', 'success')
                return redirect(url_for('dashboard'))
            else:
                error = 'Invalid credentials.'
    captcha = ''.join(random.choices(
        string.ascii_uppercase + string.digits, k=5))
    session['captcha'] = captcha
    return render_template('login.html', error=error, captcha=captcha)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    error = None
    success = None
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        mobile = request.form.get('mobile', '').strip()
        email = request.form.get('email', '').strip() or None
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        captcha_input = request.form.get('captcha', '').strip()
        if captcha_input.upper() != session.get('captcha', '').upper():
            error = 'Invalid CAPTCHA. Please try again.'
        elif not full_name or not mobile or not password:
            error = 'Full name, mobile, and password are required.'
        elif len(password) < 6:
            error = 'Password must be at least 6 characters.'
        elif password != confirm_password:
            error = 'Passwords do not match.'
        else:
            conn = get_db()
            emp_id = generate_employee_id(conn)
            try:
                conn.execute(
                    'INSERT INTO users (employee_id, full_name, email, mobile, password_hash, role) VALUES (?,?,?,?,?,?)',
                    (emp_id, full_name, email, mobile, hash_password(password), 'caller'))
                conn.commit()
                success = (f'Account created! Your Employee ID is <strong>{emp_id}</strong>. '
                           f'Role: <strong>Caller</strong>. Contact Admin to upgrade permissions.')
            except sqlite3.IntegrityError:
                error = 'An account with this mobile or email already exists.'
            finally:
                conn.close()
    captcha = ''.join(random.choices(
        string.ascii_uppercase + string.digits, k=5))
    session['captcha'] = captcha
    return render_template('registartion.html', error=error, success=success, captcha=captcha)


@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    step = request.args.get('step', '1')
    if request.method == 'POST':
        if step == '1':
            identifier = request.form.get('identifier', '').strip()
            conn = get_db()
            user = conn.execute('SELECT * FROM users WHERE email=? OR mobile=?',
                                (identifier, identifier)).fetchone()
            conn.close()
            if user:
                otp = ''.join(random.choices(string.digits, k=6))
                session['reset_otp'] = otp
                session['reset_user_id'] = user['id']
                flash(f'OTP sent! (Demo OTP: {otp})', 'info')
                return redirect(url_for('forgot_password', step='2'))
            else:
                flash('No account found.', 'danger')
        elif step == '2':
            if request.form.get('otp', '').strip() == session.get('reset_otp', ''):
                return redirect(url_for('forgot_password', step='3'))
            else:
                flash('Invalid OTP.', 'danger')
        elif step == '3':
            new_pass = request.form.get('new_password', '')
            confirm = request.form.get('confirm_password', '')
            if new_pass != confirm:
                flash('Passwords do not match.', 'danger')
            elif len(new_pass) < 6:
                flash('Minimum 6 characters.', 'danger')
            else:
                conn = get_db()
                conn.execute('UPDATE users SET password_hash=? WHERE id=?',
                             (hash_password(new_pass), session.get('reset_user_id')))
                conn.commit()
                conn.close()
                session.pop('reset_otp', None)
                session.pop('reset_user_id', None)
                flash('Password reset! Please login.', 'success')
                return redirect(url_for('login'))
    return render_template('forget_password.html', step=step)


# ─────────────────────────────────────────────
# ROUTES: DASHBOARD
# ─────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    user_id = session['user_id']
    role = session['role']
    today = datetime.now().strftime('%Y-%m-%d')

    if role == 'caller':
        total_leads = conn.execute(
            'SELECT COUNT(*) FROM leads WHERE caller_id=?', (user_id,)).fetchone()[0]
        calls_today = conn.execute(
            'SELECT COUNT(*) FROM call_logs WHERE caller_id=? AND DATE(call_date)=?', (user_id, today)).fetchone()[0]
        interested = conn.execute(
            'SELECT COUNT(*) FROM leads WHERE caller_id=? AND status="Interested"', (user_id,)).fetchone()[0]
        followups = conn.execute(
            'SELECT COUNT(*) FROM follow_ups WHERE caller_id=? AND DATE(follow_up_date)=? AND status="Pending"', (user_id, today)).fetchone()[0]
        recent_leads = conn.execute('''SELECT l.*, u.full_name as caller_name, cp.lead_category, cp.lead_score
            FROM leads l LEFT JOIN users u ON l.caller_id=u.id
            LEFT JOIN customer_profiles cp ON l.id=cp.lead_id
            WHERE l.caller_id=? ORDER BY l.created_at DESC LIMIT 10''', (user_id,)).fetchall()
    else:
        total_leads = conn.execute('SELECT COUNT(*) FROM leads').fetchone()[0]
        calls_today = conn.execute(
            'SELECT COUNT(*) FROM call_logs WHERE DATE(call_date)=?', (today,)).fetchone()[0]
        interested = conn.execute(
            'SELECT COUNT(*) FROM leads WHERE status="Interested"').fetchone()[0]
        followups = conn.execute(
            'SELECT COUNT(*) FROM follow_ups WHERE DATE(follow_up_date)=? AND status="Pending"', (today,)).fetchone()[0]
        recent_leads = conn.execute('''SELECT l.*, u.full_name as caller_name, cp.lead_category, cp.lead_score
            FROM leads l LEFT JOIN users u ON l.caller_id=u.id
            LEFT JOIN customer_profiles cp ON l.id=cp.lead_id
            ORDER BY l.created_at DESC LIMIT 10''').fetchall()

    hot_count = conn.execute(
        "SELECT COUNT(*) FROM customer_profiles WHERE lead_category='Hot'").fetchone()[0]
    warm_count = conn.execute(
        "SELECT COUNT(*) FROM customer_profiles WHERE lead_category='Warm'").fetchone()[0]
    cold_count = conn.execute(
        "SELECT COUNT(*) FROM customer_profiles WHERE lead_category='Cold'").fetchone()[0]
    conn.close()
    return render_template('dashboard.html', total_leads=total_leads, calls_today=calls_today,
                           interested=interested, followups=followups, recent_leads=recent_leads,
                           hot_count=hot_count, warm_count=warm_count, cold_count=cold_count)


# ─────────────────────────────────────────────
# ROUTES: LEADS
# ─────────────────────────────────────────────

@app.route('/leads')
@login_required
def leads():
    conn = get_db()
    role = session['role']
    user_id = session['user_id']
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    category_filter = request.args.get('category', '')
    followup_filter = request.args.get('followup', '')

    query = '''SELECT l.*, u.full_name as caller_name,
        cp.lead_category, cp.lead_score, cp.budget_range, cp.investment_timeline
        FROM leads l LEFT JOIN users u ON l.caller_id=u.id
        LEFT JOIN customer_profiles cp ON l.id=cp.lead_id WHERE 1=1'''
    params = []
    if role == 'caller':
        query += ' AND l.caller_id=?'
        params.append(user_id)
    if search:
        query += ' AND (l.name LIKE ? OR l.mobile LIKE ? OR l.email LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    if status_filter:
        query += ' AND l.status=?'
        params.append(status_filter)
    if category_filter:
        query += ' AND cp.lead_category=?'
        params.append(category_filter)
    if followup_filter == 'today':
        today = datetime.now().strftime('%Y-%m-%d')
        query += ''' AND l.id IN (SELECT lead_id FROM follow_ups WHERE DATE(follow_up_date)=? AND status='Pending')'''
        params.append(today)

    query += ' ORDER BY l.created_at DESC'

    all_leads = conn.execute(query, params).fetchall()
    callers = conn.execute(
        "SELECT id, full_name FROM users WHERE role='caller' AND is_active=1").fetchall()
    conn.close()
    return render_template('leads.html', leads=all_leads, callers=callers,
                           search=search, status_filter=status_filter, category_filter=category_filter,
                           followup_filter=followup_filter)


@app.route('/leads/add', methods=['GET', 'POST'])
@login_required
def add_lead():
    """
    New advanced lead flow:
    Step 1: Call Done? Yes/No
    Step 2 (if Yes): Connected? Yes/No + reason
    Step 3 (if Connected Yes): Interested/Not Interested
    If Interested → redirect to profile_form
    """
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        mobile = request.form.get('mobile', '').strip()
        flow_step = request.form.get('flow_step', 'save')

        if not name or not mobile:
            flash('Name and Mobile are required.', 'danger')
            return redirect(url_for('add_lead'))

        if flow_step == 'save':
            # Simple save — Not Called / Not Connected / Not Interested
            status = request.form.get('status', 'New')
            not_connected_reason = request.form.get('not_connected_reason', '')
            conn = get_db()
            cursor = conn.execute(
                'INSERT INTO leads (name, mobile, status, not_connected_reason, caller_id) VALUES (?,?,?,?,?)',
                (name, mobile, status, not_connected_reason, session['user_id']))
            lead_id = cursor.lastrowid
            conn.commit()
            conn.close()
            flash('Lead saved successfully!', 'success')
            return redirect(url_for('leads'))

        elif flow_step == 'interested':
            # Interested — save with Interested status then redirect to profile
            conn = get_db()
            cursor = conn.execute(
                'INSERT INTO leads (name, mobile, status, caller_id) VALUES (?,?,?,?)',
                (name, mobile, 'Interested', session['user_id']))
            lead_id = cursor.lastrowid
            conn.commit()
            conn.close()
            flash('Lead marked as Interested! Please fill in the profile.', 'success')
            return redirect(url_for('profile_form', lead_id=lead_id, step=1))

    return render_template('add_lead.html')


@app.route('/leads/<int:lead_id>/view')
@login_required
def view_lead(lead_id):
    conn = get_db()
    lead = conn.execute(
        'SELECT l.*, u.full_name as caller_name FROM leads l LEFT JOIN users u ON l.caller_id=u.id WHERE l.id=?', (lead_id,)).fetchone()
    profile = conn.execute(
        'SELECT * FROM customer_profiles WHERE lead_id=?', (lead_id,)).fetchone()
    call_logs = conn.execute('''SELECT cl.*, u.full_name as caller_name
        FROM call_logs cl LEFT JOIN users u ON cl.caller_id=u.id
        WHERE cl.lead_id=? ORDER BY cl.call_date DESC''', (lead_id,)).fetchall()
    followups = conn.execute('''SELECT fu.*, u.full_name as caller_name
        FROM follow_ups fu LEFT JOIN users u ON fu.caller_id=u.id
        WHERE fu.lead_id=? ORDER BY fu.follow_up_date''', (lead_id,)).fetchall()
    callers = conn.execute(
        "SELECT id, full_name FROM users WHERE role='caller' AND is_active=1").fetchall()
    conn.close()
    if not lead:
        flash('Lead not found.', 'danger')
        return redirect(url_for('leads'))
    return render_template('view_lead.html', lead=lead, profile=profile,
                           call_logs=call_logs, followups=followups, callers=callers)


@app.route('/leads/<int:lead_id>/update-status', methods=['POST'])
@login_required
def update_lead_status(lead_id):
    new_status = request.form.get('status')
    conn = get_db()
    conn.execute('UPDATE leads SET status=?, updated_at=? WHERE id=?',
                 (new_status, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), lead_id))
    conn.commit()
    conn.close()
    if new_status == 'Interested':
        return redirect(url_for('profile_form', lead_id=lead_id, step=1))
    flash('Lead status updated.', 'success')
    return redirect(url_for('view_lead', lead_id=lead_id))


@app.route('/leads/<int:lead_id>/log-call', methods=['POST'])
@login_required
def log_call(lead_id):
    conn = get_db()
    conn.execute('INSERT INTO call_logs (lead_id, caller_id, notes, duration_minutes) VALUES (?,?,?,?)',
                 (lead_id, session['user_id'], request.form.get('notes', ''), request.form.get('duration', 0)))
    conn.commit()
    conn.close()
    flash('Call logged!', 'success')
    return redirect(url_for('view_lead', lead_id=lead_id))


@app.route('/leads/<int:lead_id>/add-followup', methods=['POST'])
@login_required
def add_followup(lead_id):
    follow_up_date = request.form.get('follow_up_date', '').strip()
    notes = request.form.get('notes', '')

    # Backend validation: follow-up date must be today or future
    if follow_up_date:
        try:
            fu_dt = datetime.strptime(follow_up_date, '%Y-%m-%d').date()
            today = datetime.now().date()
            if fu_dt < today:
                flash('Follow-up date cannot be in the past.', 'danger')
                return redirect(url_for('view_lead', lead_id=lead_id))
        except ValueError:
            flash('Invalid follow-up date format.', 'danger')
            return redirect(url_for('view_lead', lead_id=lead_id))

    conn = get_db()
    conn.execute('INSERT INTO follow_ups (lead_id, caller_id, follow_up_date, notes) VALUES (?,?,?,?)',
                 (lead_id, session['user_id'], follow_up_date, notes))
    conn.commit()
    conn.close()
    flash('Follow-up scheduled!', 'success')
    return redirect(url_for('view_lead', lead_id=lead_id))


# ─────────────────────────────────────────────
# ROUTES: MULTI-STEP PROFILE FORM
# ─────────────────────────────────────────────

@app.route('/leads/<int:lead_id>/profile', methods=['GET', 'POST'])
@login_required
def profile_form(lead_id):
    step = int(request.args.get('step', 1))
    conn = get_db()
    lead = conn.execute('SELECT * FROM leads WHERE id=?',
                        (lead_id,)).fetchone()
    if not lead:
        conn.close()
        flash('Lead not found.', 'danger')
        return redirect(url_for('leads'))

    if request.method == 'POST':
        existing = conn.execute(
            'SELECT * FROM customer_profiles WHERE lead_id=?', (lead_id,)).fetchone()
        if not existing:
            conn.execute(
                'INSERT INTO customer_profiles (lead_id) VALUES (?)', (lead_id,))
            conn.commit()

        if step == 1:
            conn.execute('''UPDATE leads SET whatsapp=?, email=?, city=?, state=?, age_group=?, updated_at=? WHERE id=?''',
                         (request.form.get('whatsapp', ''), request.form.get('email', ''),
                          request.form.get(
                              'city', ''), request.form.get('state', ''),
                          request.form.get('age_group', ''),
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S'), lead_id))
        elif step == 2:
            conn.execute('''UPDATE customer_profiles SET occupation=?, company_name=?,
                monthly_income=?, investment_experience=? WHERE lead_id=?''',
                         (request.form.get('occupation', ''), request.form.get('company_name', ''),
                          request.form.get('monthly_income', ''), request.form.get(
                             'investment_experience', ''),
                             lead_id))
        elif step == 3:
            payment_pref = request.form.get('payment_preference', '')
            loan_bank = request.form.get(
                'loan_bank', '') if payment_pref == 'Loan' else ''
            loan_amount = request.form.get(
                'loan_amount', '') if payment_pref == 'Loan' else ''
            loan_tenure_raw = request.form.get(
                'loan_tenure_months', '') if payment_pref == 'Loan' else ''
            loan_tenure = int(
                loan_tenure_raw) if loan_tenure_raw and loan_tenure_raw.isdigit() else None
            conn.execute('''UPDATE customer_profiles SET budget_range=?, payment_preference=?,
                loan_bank=?, loan_amount=?, loan_tenure_months=? WHERE lead_id=?''',
                         (request.form.get('budget_range', ''), payment_pref,
                          loan_bank, loan_amount, loan_tenure, lead_id))
        elif step == 4:
            conn.execute('UPDATE customer_profiles SET purpose=?, holding_period=? WHERE lead_id=?',
                         (request.form.get('purpose', ''), request.form.get('holding_period', ''), lead_id))
        elif step == 5:
            drivers = ','.join(request.form.getlist('interest_drivers'))
            conn.execute('''UPDATE customer_profiles SET dholera_aware=?, awareness_source=?,
                interest_drivers=? WHERE lead_id=?''',
                         (request.form.get('dholera_aware', ''), request.form.get('awareness_source', ''),
                          drivers, lead_id))
        elif step == 6:
            conn.execute('UPDATE customer_profiles SET plot_size=?, project_type=? WHERE lead_id=?',
                         (request.form.get('plot_size', ''), request.form.get('project_type', ''), lead_id))
        elif step == 7:
            conn.execute('UPDATE customer_profiles SET risk_appetite=? WHERE lead_id=?',
                         (request.form.get('risk_appetite', ''), lead_id))
        elif step == 8:
            conn.execute('UPDATE customer_profiles SET investment_timeline=? WHERE lead_id=?',
                         (request.form.get('investment_timeline', ''), lead_id))
        elif step == 9:
            # Backend validation for follow-up date
            follow_up_date = request.form.get('follow_up_date', '')
            if follow_up_date:
                try:
                    fu_dt = datetime.strptime(
                        follow_up_date, '%Y-%m-%d').date()
                    today = datetime.now().date()
                    if fu_dt < today:
                        flash(
                            'Follow-up date cannot be in the past. Please choose today or a future date.', 'danger')
                        conn.close()
                        return redirect(url_for('profile_form', lead_id=lead_id, step=9))
                except ValueError:
                    flash('Invalid follow-up date.', 'danger')
                    conn.close()
                    return redirect(url_for('profile_form', lead_id=lead_id, step=9))

            conn.execute('''UPDATE customer_profiles SET lead_source=?, assigned_executive=?,
                follow_up_date=?, remarks=? WHERE lead_id=?''',
                         (request.form.get('lead_source', ''), request.form.get('assigned_executive', ''),
                          follow_up_date, request.form.get('remarks', ''), lead_id))
            conn.execute("UPDATE leads SET status='Profiled', updated_at=? WHERE id=?",
                         (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), lead_id))

        # Always recalculate score
        recalculate_and_save_score(conn, lead_id)
        conn.commit()

        if step == 9:
            conn.close()
            flash('Profile saved! Lead score updated.', 'success')
            return redirect(url_for('view_lead', lead_id=lead_id))

        conn.close()
        return redirect(url_for('profile_form', lead_id=lead_id, step=step + 1))

    profile = conn.execute(
        'SELECT * FROM customer_profiles WHERE lead_id=?', (lead_id,)).fetchone()
    conn.close()
    return render_template('profile_form.html', lead=lead, step=step, profile=profile, total_steps=9)


# ─────────────────────────────────────────────
# ROUTES: ADMIN
# ─────────────────────────────────────────────

@app.route('/admin')
@role_required('admin')
def admin_panel():
    conn = get_db()
    users = conn.execute(
        'SELECT * FROM users ORDER BY role, full_name').fetchall()
    total_leads = conn.execute('SELECT COUNT(*) FROM leads').fetchone()[0]
    total_users = conn.execute(
        'SELECT COUNT(*) FROM users WHERE is_active=1').fetchone()[0]
    total_callers = conn.execute(
        "SELECT COUNT(*) FROM users WHERE role='caller' AND is_active=1").fetchone()[0]
    hot_leads = conn.execute(
        "SELECT COUNT(*) FROM customer_profiles WHERE lead_category='Hot'").fetchone()[0]
    performance = conn.execute('''SELECT u.full_name, u.employee_id,
        COUNT(l.id) as total_leads,
        SUM(CASE WHEN l.status="Interested" THEN 1 ELSE 0 END) as interested,
        SUM(CASE WHEN l.status="Converted" THEN 1 ELSE 0 END) as converted,
        COUNT(cl.id) as total_calls
        FROM users u
        LEFT JOIN leads l ON u.id=l.caller_id
        LEFT JOIN call_logs cl ON u.id=cl.caller_id
        WHERE u.role="caller"
        GROUP BY u.id ORDER BY total_leads DESC''').fetchall()
    conn.close()
    return render_template('admin.html', users=users, total_leads=total_leads,
                           total_users=total_users, total_callers=total_callers,
                           hot_leads=hot_leads, performance=performance)


@app.route('/admin/users/add', methods=['GET', 'POST'])
@role_required('admin')
def add_user():
    if request.method == 'POST':
        emp_id = request.form.get('employee_id', '').strip()
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip()
        mobile = request.form.get('mobile', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'caller')
        if not all([emp_id, full_name, mobile, password]):
            flash('All required fields must be filled.', 'danger')
            return redirect(url_for('add_user'))
        conn = get_db()
        try:
            conn.execute('INSERT INTO users (employee_id, full_name, email, mobile, password_hash, role) VALUES (?,?,?,?,?,?)',
                         (emp_id, full_name, email, mobile, hash_password(password), role))
            conn.commit()
            flash(f'User {full_name} created!', 'success')
        except sqlite3.IntegrityError:
            flash('Employee ID, Email or Mobile already exists.', 'danger')
        finally:
            conn.close()
        return redirect(url_for('admin_panel'))
    return render_template('add_user.html')


@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@role_required('admin')
def toggle_user(user_id):
    conn = get_db()
    user = conn.execute(
        'SELECT is_active FROM users WHERE id=?', (user_id,)).fetchone()
    if user:
        conn.execute('UPDATE users SET is_active=? WHERE id=?',
                     (0 if user['is_active'] else 1, user_id))
        conn.commit()
        flash('User status updated.', 'success')
    conn.close()
    return redirect(url_for('admin_panel'))


@app.route('/admin/leads/assign', methods=['POST'])
@role_required('admin', 'manager')
def assign_lead():
    lead_id = request.form.get('lead_id')
    caller_id = request.form.get('caller_id')
    conn = get_db()
    conn.execute('UPDATE leads SET caller_id=?, updated_at=? WHERE id=?',
                 (caller_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), lead_id))
    conn.commit()
    conn.close()
    flash('Lead assigned!', 'success')
    return redirect(url_for('leads'))


# ─────────────────────────────────────────────
# ROUTES: MANAGER
# ─────────────────────────────────────────────

@app.route('/manager')
@role_required('admin', 'manager')
def manager_panel():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    callers = conn.execute('''SELECT u.*,
        COUNT(DISTINCT l.id) as total_leads,
        COUNT(DISTINCT cl.id) as calls_today,
        SUM(CASE WHEN l.status="Interested" THEN 1 ELSE 0 END) as interested
        FROM users u
        LEFT JOIN leads l ON u.id=l.caller_id
        LEFT JOIN call_logs cl ON u.id=cl.caller_id AND DATE(cl.call_date)=?
        WHERE u.role="caller" AND u.is_active=1
        GROUP BY u.id''', (today,)).fetchall()
    followups_today = conn.execute('''SELECT fu.*, l.name as lead_name, l.mobile, u.full_name as caller_name
        FROM follow_ups fu
        LEFT JOIN leads l ON fu.lead_id=l.id
        LEFT JOIN users u ON fu.caller_id=u.id
        WHERE DATE(fu.follow_up_date)=? AND fu.status="Pending"
        ORDER BY fu.follow_up_date''', (today,)).fetchall()
    conn.close()
    return render_template('manager.html', callers=callers, followups_today=followups_today)


# ─────────────────────────────────────────────
# ROUTES: EXCEL EXPORT / IMPORT (ALL FIELDS)
# ─────────────────────────────────────────────

@app.route('/export/leads')
@login_required
def export_leads():
    if not EXCEL_AVAILABLE:
        flash('Excel requires openpyxl.', 'warning')
        return redirect(url_for('leads'))

    conn = get_db()
    rows = conn.execute('''
        SELECT l.id, l.name, l.mobile, l.whatsapp, l.email, l.city, l.state,
            l.age_group, l.status, u.full_name as caller_name, l.created_at, l.updated_at,
            cp.occupation, cp.company_name, cp.monthly_income, cp.investment_experience,
            cp.budget_range, cp.payment_preference, cp.loan_bank, cp.loan_amount, cp.loan_tenure_months,
            cp.purpose, cp.holding_period,
            cp.dholera_aware, cp.awareness_source, cp.interest_drivers,
            cp.plot_size, cp.project_type,
            cp.risk_appetite, cp.investment_timeline,
            cp.lead_source, cp.assigned_executive, cp.follow_up_date, cp.remarks,
            cp.lead_score, cp.lead_category
        FROM leads l
        LEFT JOIN users u ON l.caller_id=u.id
        LEFT JOIN customer_profiles cp ON l.id=cp.lead_id
        ORDER BY l.created_at DESC
    ''').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "All Leads"

    headers = [
        'Lead ID', 'Full Name', 'Mobile', 'WhatsApp', 'Email', 'City', 'State',
        'Age Group', 'Status', 'Assigned Caller', 'Created At', 'Updated At',
        'Occupation', 'Company/Business', 'Monthly Income', 'Inv. Experience',
        'Budget Range', 'Payment Preference', 'Loan Bank', 'Loan Amount', 'Loan Tenure (Months)',
        'Investment Purpose', 'Holding Period',
        'Dholera Aware', 'Awareness Source', 'Interest Drivers',
        'Preferred Plot Size', 'Project Type',
        'Risk Appetite', 'Investment Timeline',
        'Lead Source', 'Assigned Executive', 'Follow-up Date', 'Remarks',
        'Lead Score (/13)', 'Lead Category'
    ]

    header_fill = PatternFill("solid", fgColor="1C3A1C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    cat_fills = {
        'Hot': PatternFill("solid", fgColor="FFCCCC"),
        'Warm': PatternFill("solid", fgColor="FFF3CC"),
        'Cold': PatternFill("solid", fgColor="E8F5FF"),
    }
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(list(row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cat = row['lead_category']
            if cat in cat_fills:
                cell.fill = cat_fills[cat]

    for col in ws.columns:
        max_len = max((len(str(cell.value or ''))
                      for cell in col), default=8) + 3
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 35)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"krishna_leads_full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/template')
@login_required
def download_template():
    if not EXCEL_AVAILABLE:
        flash('Excel requires openpyxl.', 'warning')
        return redirect(url_for('leads'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    headers = [
        'Name*', 'Mobile*', 'WhatsApp', 'Email', 'City', 'State', 'Age Group',
        'Occupation', 'Company Name', 'Monthly Income', 'Investment Experience',
        'Budget Range', 'Payment Preference', 'Loan Bank', 'Loan Amount (₹)', 'Loan Tenure (Months)',
        'Investment Purpose', 'Holding Period',
        'Dholera Aware (Yes/No)', 'Awareness Source', 'Interest Drivers',
        'Plot Size', 'Project Type', 'Risk Appetite', 'Investment Timeline',
        'Lead Source', 'Remarks'
    ]
    header_fill = PatternFill("solid", fgColor="C8871A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    sample = [
        'Ramesh Patel', '9876543210', '9876543210', 'ramesh@email.com',
        'Ahmedabad', 'Gujarat', '30–45',
        'Business Owner', 'Patel Traders', '₹1L–₹3L', 'Experienced Investor',
        '₹25–50 Lakhs', 'Loan', 'SBI', '2500000', '120',
        'Long-Term Growth', '5–10 Years',
        'Yes', 'Social Media', 'Price Appreciation,Smart City Development',
        '150–250 sq. yds', 'NA-NOC Clear', 'Moderate', 'Within 1 Month',
        'Reference', 'Very interested – follow up next week'
    ]
    ws.append(sample)
    note_row = ws.max_row + 1
    note_cell = ws.cell(row=note_row, column=1,
                        value='* Required. Row 2 is a sample – delete it before importing.')
    note_cell.font = Font(italic=True, color="888888")
    try:
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=8)
    except Exception:
        pass

    for col in ws.columns:
        max_len = max((len(str(cell.value or ''))
                      for cell in col), default=8) + 3
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='krishna_import_template_full.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/import/leads', methods=['POST'])
@login_required
def import_leads():
    if not EXCEL_AVAILABLE:
        flash('Excel requires openpyxl.', 'warning')
        return redirect(url_for('leads'))
    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('Please upload a valid .xlsx file.', 'danger')
        return redirect(url_for('leads'))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        skipped = 0
        conn = get_db()
        SKIP_PHRASES = {'name', 'name*', 'sample name',
                        '* required fields', '* required. row 2 is a sample – delete it before importing.'}

        def gv(r, i, default=''):
            val = r[i] if len(r) > i else None
            s = str(val).strip() if val is not None else ''
            return s if s and s.lower() not in ('none', '') else default

        for row in ws.iter_rows(min_row=2, values_only=True):
            name_str = gv(row, 0)
            mobile_str = gv(row, 1)
            if not name_str or not mobile_str or name_str.lower() in SKIP_PHRASES:
                skipped += 1
                continue
            try:
                cursor = conn.execute('''INSERT INTO leads (name, mobile, whatsapp, email, city, state,
                    age_group, caller_id) VALUES (?,?,?,?,?,?,?,?)''',
                                      (name_str, mobile_str, gv(row, 2), gv(row, 3), gv(row, 4),
                                       gv(row, 5), gv(row, 6), session['user_id']))
                lead_id = cursor.lastrowid

                occ = gv(row, 7)
                company = gv(row, 8)
                income = gv(row, 9)
                exp = gv(row, 10)
                budget = gv(row, 11)
                payment = gv(row, 12)
                l_bank = gv(row, 13)
                l_amt = gv(row, 14)
                l_ten_raw = gv(row, 15)
                l_ten = int(
                    l_ten_raw) if l_ten_raw and l_ten_raw.isdigit() else None
                purpose = gv(row, 16)
                holding = gv(row, 17)
                aware = gv(row, 18)
                src = gv(row, 19)
                drivers = gv(row, 20)
                plot = gv(row, 21)
                proj = gv(row, 22)
                risk = gv(row, 23)
                timeline = gv(row, 24)
                lead_src = gv(row, 25)
                remarks = gv(row, 26)

                if any([occ, budget, payment, purpose, timeline]):
                    conn.execute('''INSERT INTO customer_profiles
                        (lead_id, occupation, company_name, monthly_income, investment_experience,
                         budget_range, payment_preference, loan_bank, loan_amount, loan_tenure_months,
                         purpose, holding_period, dholera_aware, awareness_source, interest_drivers,
                         plot_size, project_type, risk_appetite, investment_timeline, lead_source, remarks)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                                 (lead_id, occ, company, income, exp, budget, payment,
                                  l_bank, l_amt, l_ten, purpose, holding, aware, src, drivers,
                                  plot, proj, risk, timeline, lead_src, remarks))
                    recalculate_and_save_score(conn, lead_id)

                imported += 1
            except sqlite3.IntegrityError:
                skipped += 1

        conn.commit()
        conn.close()
        msg = f'Imported {imported} lead(s)!'
        if skipped:
            msg += f' ({skipped} skipped)'
        flash(msg, 'success')
    except Exception as e:
        flash(f'Import error: {str(e)}', 'danger')
    return redirect(url_for('leads'))


# ─────────────────────────────────────────────
# ROUTES: REPORTS
# ─────────────────────────────────────────────

@app.route('/reports')
@role_required('admin', 'manager')
def reports():
    conn = get_db()
    status_data = conn.execute(
        'SELECT status, COUNT(*) as count FROM leads GROUP BY status').fetchall()
    category_data = conn.execute(
        'SELECT lead_category, COUNT(*) as count FROM customer_profiles GROUP BY lead_category').fetchall()
    caller_perf = conn.execute('''SELECT u.full_name, COUNT(l.id) as leads,
        SUM(CASE WHEN l.status="Converted" THEN 1 ELSE 0 END) as converted
        FROM users u LEFT JOIN leads l ON u.id=l.caller_id
        WHERE u.role="caller" GROUP BY u.id ORDER BY leads DESC''').fetchall()
    monthly = conn.execute('''SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count
        FROM leads GROUP BY month ORDER BY month DESC LIMIT 12''').fetchall()
    conn.close()
    return render_template('reports.html', status_data=status_data,
                           category_data=category_data, caller_perf=caller_perf, monthly=monthly)


@app.route('/api/lead-stats')
@login_required
def api_lead_stats():
    conn = get_db()
    data = {
        'hot': conn.execute("SELECT COUNT(*) FROM customer_profiles WHERE lead_category='Hot'").fetchone()[0],
        'warm': conn.execute("SELECT COUNT(*) FROM customer_profiles WHERE lead_category='Warm'").fetchone()[0],
        'cold': conn.execute("SELECT COUNT(*) FROM customer_profiles WHERE lead_category='Cold'").fetchone()[0],
    }
    conn.close()
    return jsonify(data)


# ─────────────────────────────────────────────
# PROFILE FORM: User profile page
# ─────────────────────────────────────────────

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?',
                        (session['user_id'],)).fetchone()
    conn.close()
    return render_template('profile_form_user.html', user=user)




if __name__ == '__main__':
    
    print("\n" + "="*55)
    print("  KRISHNA LAND DEVELOPERS - CRM SYSTEM")
    print("="*55)
    print("  Server: http://127.0.0.1:5000")
    print("  Admin   → KLD001 / Admin@123")
    print("  Manager → KLD002 / Manager@123")
    print("  Caller  → KLD003 / Caller@123")
    print("="*55 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
